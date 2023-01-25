# Created by Joshua Hardy and mmason
from genericpath import isfile
import os
import re
import xlrd
import openpyxl
import datetime
import time
import shutil
import tkinter
from tkinter import filedialog
from pathlib import Path
import csv

{}
# Query imports
import sys
sys.path.insert(1, '../Files/')
from Files import After_Repack_Queries
from Files import Alt_Loan_Queries
from Files import Atb_Fbill_3C_Queries
from Files import Budget_Queries
from Files import Daily_Queries
from Files import Day_AfterLDR
from Files import Direct_Loan
from Files import Disbursement_Queries
from Files import EndOfTerm_Queries
from Files import Mid_Repack_Queries
from Files import Monday_WeeklyQueries
from Files import Monthly_Queries
from Files import Packaging_Queries
from Files import PrePackaging_Queries
from Files import Scholarships_Queries
from Files import Second_LDR
from Files import Tsm_Queries


global test

skip_files = [".zip", ".lnk", " DL ORIG ", " ALT Loan ORIG "]
remove_files = ["FASTDVER", "FINAID_Checklist", "ussfa09", "USSFA090 Reset", "O-A", r"uosfa[0-9]+[a-zA-Z]*\.csv"]

# The date becomes the current date and is then placed in MM-DD-YY format
date = time.strftime("%x").replace("/", "-")
now = datetime.datetime.now()
last_month = now.month - 1 if now.month > 1 else 12
last_months_year = now.year - 1 if now.month == 12 else now.year
month_folder = date[:2] + "-20" + date[-2:]

# Var definitions
odd_aid_years = []
even_aid_years = []
unknown_list = []
current_aid_year = ""
folder_path = Path()
disbursement_date = datetime.datetime.min
alt_loan_flag = False
direct_loan_flag = False

# Regular Expressions
aid_year_regex = ["Aid[\s]?Y(ea)?r"]
aid_num_regex = ["[0-9]{2,4}[\s]*$"]
date_regex = ["(0*[1-9]|1[012])[-/.](0*[1-9]|[12][0-9]|3[01])[-/.](2\d{3}|\d{2})","(0*[1-9]|[12][0-9]|3[01])[-/.](0*[1-9]|1[012])[-/.](2\d{3}|\d{2})"]
instance_regex = r"[_-][0-9_-]+\."

# Directories
test_UOSFA_directory = Path("C:/Users/iessaghir/Documents/DoQueries/Destination Folders")
#test_UOSFA_directory = Path("O:/UOSFA Reports")
#test_UOSFA_directory = Path("C:/Users/JHARDY/Documents/DoQueries/Destination Folders")

UOSFA_directory = Path("O:/UOSFA Reports")

# Origination Filepaths
test_dir_orig_folder = Path('C:/Testing Bob/Direct Loans/Origination')
test_alt_orig_folder = Path('C:/Testing Bob/ALT Loans/')
dir_orig_folder = Path('O:/Systems/Direct Loans/Origination')
alt_orig_folder = Path('O:/Systems/QUERIES/ALT Loans')

# Automatically Updated Queries
query_dict = {} # Data format =  {"Query Name" : "UOSFA Folder"}
dict_path = "./Files/Query_Dictionary.csv"

# Odd year
def is_odd_year(year):
    year_int = int(year[-1])
    return year_int % 2 == 1

def is_aid_year_word(value):
    return any (re.search(regex_str, str(value), re.IGNORECASE) for regex_str in aid_year_regex)

def is_aid_year_num(value):
    # Find 2 to 4 digit number
    result = re.search(aid_num_regex[0], str(value))
    
    # Check if number is within aid year range
    if result:
        match = result.group(0)
        match_num = int(match)
        if len(match) == 4:
            curr_num = int(current_aid_year)
            return (match_num > curr_num - 5) and (match_num < curr_num + 5)
        if len(match) == 3:
            return False
        if len(match) == 2:
            curr_num = int(str(current_aid_year)[-2:]) + 2000
            match_num += 2000
            return (match_num > curr_num - 5) and (match_num < curr_num + 5)

    return result


def is_date(value):
    return any (re.search(regex_str, str(value)) for regex_str in date_regex)


# Search for and return aid year in filename
def has_aid_year(filename):
    filestring = str(filename)
    has = False
    year = "0"
    found_year = re.search("_\d\d-", filestring)
    if found_year:
        has = True
        year = "20" + found_year.group()[1:-1]
    return (has, year)


# Search for and return maximum aid year in xls file
def search_xls_file(filename):
    global folder_path
    has = False
    year = "0"

    fullpath = folder_path / filename
    workbook = xlrd.open_workbook(fullpath, logfile=open(os.devnull, 'w'))
    sheet = workbook.sheet_by_index(0)
    aid_cols = []
    aid_rows = []
    max_year = 0
    
    # Locate cells containing the word 'Aid Year'
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            value = sheet.cell_value(row, col)
            if is_aid_year_word(value):
                aid_cols.append(col)
                aid_rows.append(row)
                if is_aid_year_num(value):
                    has = True
                    year = "20" +  str(value)[-2:] 
                    workbook.release_resources()
                    return (has, year)
                elif is_date(value):
                    has = True
                    year = "20" +  str(value)[-2:] # Assumes date format w/ year at end
                    workbook.release_resources()
                    return (has, year)
    
    # Find maximum aid year in 'Aid Year' column
    for i in range(len(aid_cols)):
        aid_col = aid_cols[i]
        aid_row = aid_rows[i]
        if aid_col > -1:
            curr_row = aid_row
            while (curr_row < sheet.nrows):
                value = str(sheet.cell_value(curr_row, aid_col))
                if is_aid_year_num(value):
                    value_int = int(value[-2:])
                    if value_int > max_year:
                        max_year = value_int
                        has = True
                        year = "20" +  str(value)[-2:]
                elif is_date(value):
                    value_int = int(value[-2:])
                    if value_int > max_year:
                        max_year = value_int
                        has = True
                        year = "20" +  str(value)[-2:] # Assumes date format w/ year at end
                curr_row = curr_row + 1

    workbook.release_resources()
    return (has, year)


# Search for and return aid year in excel file
def search_excel_file(filename):
    global folder_path
    has = False
    year = "0"

    fullpath = folder_path / filename
    workbook = openpyxl.load_workbook(fullpath, True)
    sheet = workbook.active
    aid_cols = []
    aid_rows = []
    max_year = 0
    
    # Locate cells containing the word 'Aid Year'
    for row in sheet.rows:
        for cell in row:
            value = str(cell.value)
            if is_aid_year_word(value):
                aid_cols.append(cell.col_idx)
                aid_rows.append(cell.row_idx)
                if is_aid_year_num(value):
                    has = True
                    year = "20" +  str(value)[-2:]
                    workbook.close()
                    return (has, year)
                elif is_date(value):
                    has = True
                    year = "20" +  str(value)[-2:] # Assumes date format w/ year at end
                    workbook.close()
                    return (has, year)

    for i in range(len(aid_cols)):
        aid_col = aid_cols[i]
        aid_row = aid_rows[i]
        curr_row = aid_row
        while (curr_row < sheet.max_row):
            value = sheet.cell(curr_row, aid_col)
            if is_aid_year_word(value):
                if is_aid_year_num(value):
                    value_int = int(value[-2:])
                    if value_int > max_year:
                        max_year = value_int
                        has = True
                        year = "20" +  str(value)[-2:]                     
                elif is_date(value):
                    value_int = int(value[-2:])
                    if value_int > max_year:
                        max_year = value_int
                        has = True
                        year = "20" +  str(value)[-2:] # Assumes date format w/ year at end
                curr_row = curr_row + 1

    workbook.close()
    return (has, year)


# Prints list of sorted files to terminal
# **(For use when debugging)**
def output_sorted_files():
    print("Num Odds = " + str(len(odd_aid_years)))
    print("Odds: ")
    for filename in odd_aid_years:
        print("- " + filename)
    print()

    print("Num Evens = " + str(len(even_aid_years)))
    print("Evens: ")
    for filename in even_aid_years:
        print("- " + filename)
    print()

    print("Num Unknown = " + str(len(unknown_list)))
    print("Unknown: ")
    for filename in unknown_list:
        print("- " + filename)
    print()


# Renames file to ensure no duplicates at desired location
def rename_no_duplicates(folder_path, renamed):
    filepath = str(folder_path / Path(renamed))
    while(isfile(filepath)):
        paren_index = filepath.find("(")
        if paren_index > -1:
            right_paren_index = filepath.find(")")
            dup_num = filepath[paren_index + 1:right_paren_index]
            dup_num = str(int(dup_num) + 1)           
            filepath = filepath[:paren_index+1] + dup_num + filepath[right_paren_index:]
        else:
            dot_index = filepath.find(".")
            filepath = filepath[:dot_index] + " (2)" + filepath[dot_index:]
    return filepath


# Renames, copies, and moves file to desired destination
def do_query(name, renamed, legacy_archive, UOSFA_folder):

    current_filepath = str(folder_path / Path(name))

    if test:
        UOSFA_destination = str(test_UOSFA_directory / UOSFA_folder)
    else:       
        UOSFA_destination = str(UOSFA_directory / UOSFA_folder)

    if UOSFA_folder == "None":
        legacy_filepath = rename_no_duplicates(legacy_archive, renamed)
        shutil.move(current_filepath, legacy_filepath)
    else:
        legacy_filepath = rename_no_duplicates(legacy_archive, renamed)
        UOSFA_filepath = rename_no_duplicates(UOSFA_destination, renamed)
        shutil.copy(current_filepath, legacy_filepath)
        shutil.move(current_filepath, UOSFA_filepath)   
 

# Returns the corresponding archive folder associated with the given UOSFA folder
def get_unknown_archive(UOSFA_folder):
    aid_year = str(int(current_aid_year) - 1) + "-" + str(current_aid_year)
    month_folder = date[:2] + "-20" + date[-2:]

    if UOSFA_folder == "Alternative Loan Reports":
        archive = os.path.realpath('O:/Systems/QUERIES/ALT Loans/')
    elif UOSFA_folder == "Budget Reports":      
        archive = os.path.realpath(os.path.join('O:/Systems/QUERIES/Budgets', aid_year, month_folder))
    elif UOSFA_folder == "Daily Reports":
        archive = os.path.realpath(os.path.join('O:/Systems/QUERIES/Daily', aid_year, month_folder))
    elif UOSFA_folder == "Direct Loan Reports":
        archive = os.path.realpath(os.path.join('O:/Systems/Direct Loans', 'DL Pre-Outbound'))
    elif UOSFA_folder == "External Award Reports":
        archive = os.path.realpath("O:\Systems\External Awards\External Award Queries")
    elif UOSFA_folder == "Financial Aid Reports":
        archive = os.path.realpath('O:/Systems/QUERIES/')
    elif UOSFA_folder == "Monthly Reports":
        archive = os.path.realpath(os.path.join('O:/Systems/QUERIES/Monthly', month_folder))
    elif UOSFA_folder == "Packaging Reports":
        archive = os.path.realpath(os.path.join('O:/Systems/QUERIES/Packaging', aid_year, month_folder))
    elif UOSFA_folder == "Pell Reports":
        archive = os.path.realpath(os.path.join('O:\Systems\QUERIES\Pell Repackaging', aid_year))
    elif UOSFA_folder == "SAP Reports":
        archive = os.path.realpath(os.path.join('O:\Systems\QUERIES/SAP/', current_aid_year))
    elif UOSFA_folder == "Scholarship Reports":
        archive = os.path.realpath(os.path.join('O:\Systems\Scholarships', aid_year + ' Scholar\Queries'))
    elif UOSFA_folder == "Unknown Reports":
        return None
    elif UOSFA_folder == "Weekly Reports":
        archive = os.path.realpath(os.path.join('O:/Systems/QUERIES/Monday Weekly', aid_year, month_folder))
    
    if not os.path.isdir(archive):
        os.makedirs(archive)
        
    return archive


# Renames and moves file to manually selected folder
def do_query_unknown(name, renamed, destination):
    global query_dict
    

    # Add query destination to query dictionary
    load_dictionary()
    if destination != "Unknown Reports":
        cleaned = clean_filename(name)
        query_dict[cleaned] = destination
    save_dictionary()
    
    # Copy to archive and move to UOSFA folder
    current_filepath = str(folder_path / Path(name))

    if test:
        UOSFA_destination = str(test_UOSFA_directory / destination)
        archive = None
    else:       
        UOSFA_destination = str(UOSFA_directory / destination)
        archive = get_unknown_archive(destination)

    if archive is not None:
        archive = rename_no_duplicates(archive, renamed)
        shutil.copy(current_filepath, archive)

    destination_filepath = rename_no_duplicates(UOSFA_destination, renamed)
    shutil.move(current_filepath, destination_filepath)


# Rename filename to be suitable for entry into query dictionary
def clean_filename(filename):
    if filename.endswith('.csv'):
        return filename

    hay_result = has_aid_year(filename)
    if hay_result[0]:
        dot_index = filename.find(".")
        res = re.search(instance_regex, filename)
        instance_index = -1
        if res is not None:
            instance_index = res.start()
        if instance_index > -1:
            renamed = filename[:(instance_index)] + filename[dot_index:]
        else:
            renamed = filename[:(dot_index - 3)] + filename[dot_index:]
    else:
        dot_index = filename.find(".")
        res = re.search(instance_regex, filename)
        instance_index = -1
        if res is not None:
            instance_index = res.start()
        if instance_index > -1:
            renamed = filename[:(instance_index)] + filename[dot_index:]
        else:
            renamed = filename[:(dot_index)] + filename[dot_index:]

    return renamed


# Returns renamed filename with current date and aid year included        
def new_name(name, year):
    hay_result = has_aid_year(name)
    renamed = ""
    if hay_result[0]:
        dot_index = name.find(".")
        #dash_index = name.rfind("-")
        #if dash_index > -1:
        #    renamed = date + " " + name[:(dash_index - 3)] + " " + year[2:] + name[dot_index:]
        #else:
        #    renamed = date + " " + name[:(dot_index - 3)] + " " + year[2:] + name[dot_index:]
        res = re.search(instance_regex, name)
        instance_index = -1
        if res is not None:
            instance_index = res.start()
        if instance_index > -1:
            renamed = date + " " + name[:(instance_index)] + " " + year[2:] + name[dot_index:]
        else:
            renamed = date + " " + name[:(dot_index - 3)] + " " + year[2:] + name[dot_index:]

    else:
        dot_index = name.find(".")
        #dash_index = name.rfind("-")
        #if dash_index > -1:
        #    renamed = date + " " + name[:dash_index] + " " + year[2:] + name[dot_index:]
        #else:
        #    renamed = date + " " + name[:dot_index] + " " + year[2:] + name[dot_index:]
        res = re.search(instance_regex, name)
        instance_index = -1
        if res is not None:
            instance_index = res.start()
        if instance_index > -1:
            renamed = date + " " + name[:(instance_index)] + " " + year[2:] + name[dot_index:]
        else:
            renamed = date + " " + name[:(dot_index)] + " " + year[2:] + name[dot_index:]
    return renamed


# Returns renamed filename with disbursement date and aid year included   
def new_name_disb(name, year):
    hay_result = has_aid_year(name)
    renamed = ""
    if hay_result[0]:
        dot_index = name.find(".")
        #dash_index = name.rfind("-")
        #renamed = disbursement_date + " " + name[:(dash_index - 3)] + " " + year[2:] + name[dot_index:]
        res = re.search(instance_regex, name)
        instance_index = -1
        if res is not None:
            instance_index = res.start()
        if instance_index > -1:
            renamed = disbursement_date + " " + name[:(instance_index)] + " " + year[2:] + name[dot_index:]
        else:
            renamed = disbursement_date + " " + name[:(dot_index - 3)] + " " + year[2:] + name[dot_index:]
    else:
        dot_index = name.find(".")
        #dash_index = name.rfind("-")
        #renamed = disbursement_date + " " + name[:dash_index] + " " + year[2:] + name[dot_index:]
        res = re.search(instance_regex, name)
        instance_index = -1
        if res is not None:
            instance_index = res.start()
        if instance_index > -1:
            renamed = disbursement_date + " " + name[:(instance_index)] + " " + year[2:] + name[dot_index:]
        else:
            renamed = disbursement_date + " " + name[:(dot_index)] + " " + year[2:] + name[dot_index:]
    return renamed


# Move files to corresponding directory
def move_files(filename, year):
    global current_aid_year
    global odd_aid_years
    global even_aid_years
    global unknown_list
    global disbursement_date
    global alt_loan_flag
    global direct_loan_flag

    date = time.strftime("%x").replace("/", "-")
    renamed = new_name(filename, year)
    renamed_disb = new_name_disb(filename, year)
    

    info = "Empty" # Stores do_query parameters

# Daily Queries
    if info == "Empty":
        info = Daily_Queries.do_dailies(test, date, year, filename, renamed)
# Monday Weekly Queries
    if info == "Empty": 
        info = Monday_WeeklyQueries.do_monday_weeklies(test, date, year, filename, renamed)
# Budget Queries
    if info == "Empty": 
        info = Budget_Queries.do_budget_queries(test, date, year, filename, renamed)
# Packaging Queries
    if info == "Empty": 
        info = Packaging_Queries.do_packaging_queries(test, date, year, filename, renamed)
# Monthly Queries
    if info == "Empty": 
        info = Monthly_Queries.do_monthlies(test, date, current_aid_year, filename, renamed)
# Disbursement Queries
    if info == "Empty": 
        info = Disbursement_Queries.do_disb_queries(test, date, year, filename, renamed_disb)
#2nd LDR Queries
    if info == "Empty": 
        info = Second_LDR.do_2nd_ldr(test, year, filename, renamed)
# End of Term Queries
    if info == "Empty": 
        info = EndOfTerm_Queries.do_end_of_term_queries(test, date, year, filename, renamed)
# Day After LDR Queries
    if info == "Empty":
        info = Day_AfterLDR.do_day_after_ldr(test, year, filename, renamed)
# Direct Loans Pre-Outbound Queries
    if info == "Empty": 
        info = Direct_Loan.dl_pre_outbound(test, date, year, filename, renamed)
        if info != "Empty":
            direct_loan_flag = True # Notify that ORIG file must be moved
## Alternative Loan Pre-Outbound Queries
    if info == "Empty": 
        info = Alt_Loan_Queries.al_pre_outbound(test, filename, renamed)
        if info != "Empty":
            alt_loan_flag = True # Notify that ORIG file must be moved
# Pre-Repackaging Queries
    if info == "Empty": 
        info = PrePackaging_Queries.do_pre_repackaging(test, year, filename, renamed)
# Mid-Repackaging Queries
    if info == "Empty": 
        info = Mid_Repack_Queries.do_mid_repack_queries(test, year, filename, renamed)
# After Repackaging Queries
    if info == "Empty": 
        info = After_Repack_Queries.do_after_repackaging(test, year, filename, renamed)
# Daily Scholarships Queries
    if info == "Empty": 
        info = Scholarships_Queries.do_daily_scholarships(test, year, filename, renamed)
# Weekly Scholarships Queries
    if info == "Empty": 
        info = Scholarships_Queries.do_weekly_scholarships(test, year, filename, renamed)
# Budget Testing Queries
    if info == "Empty": 
        info = Budget_Queries.do_budget_test_queries(test, date, year, filename, renamed)
# ATB and 3C Queries
    if info == "Empty": 
        info = Atb_Fbill_3C_Queries.do_atb_fb_3c_queries(test, filename, renamed)
# Remove extra files 
    #if "FASTDVER" in filename or "FINAID_Checklist" in filename  or "ussfa09" in filename or "USSFA090 Reset" in filename or "O-A" in filename:
    if any (re.search(regex_str, filename) for regex_str in remove_files):
        os.remove(folder_path / Path(filename))
        print("Removed " + filename)
        if is_odd_year(year):
            odd_aid_years.remove(filename)
        else:
            even_aid_years.remove(filename)
        info = "Removed"
# Transfer Student Monitoring
    if info == "Empty": 
        info = Tsm_Queries.do_tsm_queries(test, filename, renamed)
# Files stored in .csv
    if info == "Empty":
        cleaned = clean_filename(filename)
        if cleaned in query_dict:
            val = query_dict[cleaned]
            do_query_unknown(filename, renamed, val)
            return

# Unknown File
    if info == "Empty":
        unknown_list.append(str(filename))
    elif info == "Removed":
        pass
    else:
        do_query(info[0], info[1], info[2], info[3])


# Copy origination file for direct loans
def move_direct_orig(filepath, dflt):
    if test:
        source_folder = test_dir_orig_folder
        dest_folder = test_UOSFA_directory / Path("Direct Loan Reports")
    else:
        source_folder = dir_orig_folder
        dest_folder = Path("O:/UOSFA Reports/Direct Loan Reports")

    renamed = date + " DL ORIG " + current_aid_year + ".doc"   

    if (dflt):
        source_file = source_folder / Path(date + " DL ORIG " + current_aid_year + ".doc")
        source_filex = Path(str(source_file) + "x")
        source_file2 = source_folder / Path(str(date + " DL ORIG " + current_aid_year + ".doc") + " (2)")
        source_file2x = Path(str(source_file2) + "x")
    else:
        source_file = Path(filepath)
        source_filex = Path(str(source_file) + "x")
        source_file2 = Path(str(filepath) + " (2)")
        source_file2x = Path(str(source_file2) + "x")
    
    dest_file = dest_folder / Path(renamed)
    dest_filex = Path(str(dest_file) + "x") 
    dest_file2 = dest_folder / Path(str(renamed) + " (2)")
    dest_file2x = Path(str(dest_file2) + "x")
    
    # Copy orig file
    if not isfile(dest_file) and not isfile(dest_filex):
        try:
            shutil.copy(source_file, dest_file)
        except FileNotFoundError as e:
            try:
                shutil.copy(source_filex, dest_filex)
            except FileNotFoundError as e:
                return False
    
    # Copy orig file (2)
    if not isfile(dest_file2) and not isfile(dest_file2x):
        try:
            shutil.copy(source_file2, dest_file2)
        except FileNotFoundError as e:
            try:
                shutil.copy(source_file2x, dest_file2x)
            except FileNotFoundError as e:
                pass
    return True


# Copy origination file for alt loans
def move_alt_orig(filepath, dflt):
    if test:
        source_folder = test_alt_orig_folder
        dest_folder = test_UOSFA_directory / Path("Alternative Loan Reports")
    else:
        source_folder = alt_orig_folder
        dest_folder = Path("O:/UOSFA Reports/Alternative Loan Reports")

    renamed = date + " ALT Loan ORIG " + current_aid_year + ".doc" 

    if (dflt):
        source_file = source_folder / Path(date + " ALT Loan ORIG " + current_aid_year + ".doc")
        source_filex = Path(str(source_file) + "x")
    else:
        source_file = Path(filepath)
        source_filex = Path(str(source_file) + "x")

    dest_file = dest_folder / Path(renamed)
    dest_filex = Path(str(dest_file) + "x")
    
    if not isfile(dest_file) and not isfile(dest_filex):
        try:
            shutil.copy(source_file, dest_file)
        except FileNotFoundError as e:
            try:
                shutil.copy(source_filex, dest_filex)
            except FileNotFoundError as e:
                return False
    return True


def aid_year_match(year):
    global current_aid_year
    if is_odd_year(current_aid_year):
        return is_odd_year(year)
    else:
        return not is_odd_year(year)


# Sort file into list depending on aid year
def find_aid_year(filename):
    global current_aid_year
    global odd_aid_years
    global even_aid_years

    filestring = str(filename)
    if any(word in filestring for word in skip_files):
        return

    file_year = has_aid_year(filestring)
    if file_year[0]:
        if is_odd_year(file_year[1]):
            odd_aid_years.append(str(filestring))
            move_files(filestring, file_year[1])
        else:
            even_aid_years.append(str(filestring))
            move_files(filestring, file_year[1])
    elif filestring.lower().endswith("xls"):
        file_year = search_xls_file(filename)
        if file_year[0]:
            if is_odd_year(file_year[1]):
                odd_aid_years.append(str(filestring))
                move_files(filestring, file_year[1])
            else:
                even_aid_years.append(str(filestring))
                move_files(filestring, file_year[1])
        else: # Default for spreadsheets is current year
            if is_odd_year(current_aid_year):
                odd_aid_years.append(str(filestring))
                move_files(filestring, current_aid_year)
            else:
                even_aid_years.append(str(filestring))
                move_files(filestring, current_aid_year)
    elif filestring.lower().endswith(('xlsx', 'xlsm', 'xltx', 'xltm')):
        file_year = search_excel_file(filename) # file_year: (bool has_year, str year)
        if file_year[0]:
            if is_odd_year(file_year[1]):
                odd_aid_years.append(str(filestring))
                move_files(filestring, file_year[1])
            else:
                even_aid_years.append(str(filestring))
                move_files(filestring, file_year[1])
        else: # Default for spreadsheets is current year
            if is_odd_year(current_aid_year):
                odd_aid_years.append(str(filestring))
                move_files(filestring, current_aid_year)
            else:
                even_aid_years.append(str(filestring))
                move_files(filestring, current_aid_year)
    else:
        # Default is current year
        if is_odd_year(current_aid_year):
            odd_aid_years.append(str(filestring))
            move_files(filestring, current_aid_year)
        else:
            even_aid_years.append(str(filestring))
            move_files(filestring, current_aid_year)

    
# Sort all files in directory into odd or even aid years
def sort_files():
    global folder_path
    global test
    global unknown_list

    # Old Version of File Select
    #folder_path = Path(os.getcwd())
    #for filename in os.listdir("."):
    #    pFilename = Path(filename)
    #    find_aid_year(pFilename)

    root = tkinter.Tk()    
    root.withdraw()
    directory = filedialog.askdirectory()
    root.destroy()
    if directory == "":
        return
    folder_path = directory
    files = [filepath for filepath in os.listdir(directory) if os.path.isfile(Path(directory) / Path(filepath))]
    for filename in files:
        pFilename = Path(filename)
        find_aid_year(pFilename)   

# Saves the Query Dictionary to a .csv file
def save_dictionary():
    with open(dict_path, "w", newline="\n") as data:
        w = csv.writer(data)
        for key, value in query_dict.items():
            w.writerow([key, value])
    print("Dictionary Saved")

# Loads the Query Dictionary from a .csv file
def load_dictionary():
    global query_dict
    with open(dict_path) as data:
        reader = csv.reader(data)
        for rows in reader:
            if rows:
                key = rows[0]
                value = rows[1]
                query_dict[key] = value
    print("Dictionary Loaded")

def test_save():
    testdict = {"query": "folder", "UFAA": "Budget Reports"}
    with open(dict_path, "w", newline="\n") as data:
        w = csv.writer(data)
        for key, value in testdict.items():
            w.writerow([key, value])
    print("Test Save")

def test_load():
    testdict = {}
    with open(dict_path) as data:
        reader = csv.reader(data)
        for rows in reader:
            key = rows[0]
            value = rows[1]
            testdict[key] = value
    print("Test Load")


# User Input - Initialize Aid year and disbursement date
def initialize(year, is_test):
    global current_aid_year
    global STerm
    global disbursement_date
    global alt_loan_flag
    global direct_loan_flag
    global folder_path
    global test
    global odd_aid_years
    global even_aid_years
    global unknown_list

    odd_aid_years = []
    even_aid_years = []
    unknown_list = []

    test = is_test

    folder_path = Path()

    alt_loan_flag = False
    direct_loan_flag = False


    current_aid_year = year
    today = datetime.date.today()
    if today.weekday() == 0:
        disbursement_date = today - datetime.timedelta(days = 3)
        disbursement_date = disbursement_date.strftime("%m-%d-%y")
    else:
        disbursement_date = today - datetime.timedelta(days = 1)
        disbursement_date = disbursement_date.strftime("%m-%d-%y")
    if test:
        print("Disbursement Date: " + disbursement_date)

    load_dictionary()

def run(year, is_test):
    if is_test:
        initialize(year, is_test)
        sort_files()
        output_sorted_files()
        print("Done")
        
    else:
        initialize(year, is_test)
        sort_files()
        output_sorted_files()
        print("Done")

    return (direct_loan_flag, alt_loan_flag, unknown_list)
        



