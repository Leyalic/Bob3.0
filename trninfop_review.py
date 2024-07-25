import os
import shutil
import datetime
from openpyxl import Workbook, load_workbook

def copy_text_to_excel(source_folder, excel_file):
    today = datetime.date.today()
    search_name = "trninfop"
    output_sheet_name = "File Data"

    # Create/Open Excel Workbook
    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)
        sheet = workbook[output_sheet_name]
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = output_sheet_name

    # Find files created today with the specified name
    for filename in os.listdir(source_folder):
        filepath = os.path.join(source_folder, filename)
        if (
            os.path.isfile(filepath)
            and filename.startswith(search_name)
            and datetime.date.fromtimestamp(os.path.getctime(filepath)) == today
        ):
            # Read lines from the file
            with open(filepath, 'r') as file:
                file_lines = file.readlines()

            # Paste lines to successive rows in column A
            for i, line in enumerate(file_lines, start=1):
                sheet.cell(row=i, column=1, value=line.strip())

    # Save the new Excel file with the prefixed date in the same directory as source_folder_path
    date_format = today.strftime('%m-%d-%Y')
    output_file_name = f"{date_format} {os.path.basename(excel_file)}"
    output_file_path = os.path.join(source_folder, output_file_name)
    workbook.save(output_file_path)

    # Save a copy of the new Excel file to "O:\UOSFA Reports\Daily Reports"
    destination_folder = r"O:\UOSFA Reports\Daily Reports"
    shutil.copy(output_file_path, destination_folder)

if __name__ == "__main__":
    # Set the source folder path to "H:\FA\T4WAN\data"
    source_folder_path = r"H:\FA\T4WAN\data"
    
    # Set the Excel file path to "O:\Systems\Documentation\TRNINFOP Review - Copy.xlsx"
    excel_file_path = r"O:\Systems\Documentation\NSLDS Error Review.xlsx"

    copy_text_to_excel(source_folder_path, excel_file_path)
